import requests
import sys
from datetime import datetime, timedelta, timezone
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

def fetch_all_anime_entries(username):
    base_url = f'https://myanimelist.net/animelist/{username}/load.json'
    status = 7  # Status for 'completed' or other desired status
    offset = 0
    limit = 299  # Adjust if the endpoint returns a different number of results per page
    all_anime = []

    while True:
        url = f'{base_url}?status={status}&offset={offset}'
        response = requests.get(url)

        if response.status_code == 200:
            data = response.json()
            if not data:
                # No more data to fetch
                break

            # Filter out entries with status = 6
            filtered_data = [anime for anime in data if anime.get('status') != 6]
            all_anime.extend(filtered_data)
            offset += limit  # Move to the next page
        else:
            print(f"Failed to fetch data. HTTP Status code: {response.status_code}")
            break

    return all_anime

def filter_anime_by_date(anime_list, months):
    filtered_anime = []
    cutoff_date = datetime.now(timezone.utc) - timedelta(days=months*30)

    for anime in anime_list:
        updated_at = anime.get('updated_at')
        if updated_at:
            updated_at_date = datetime.fromtimestamp(int(updated_at), tz=timezone.utc)
            if updated_at_date >= cutoff_date:
                filtered_anime.append(anime)

    return filtered_anime

def print_anime_titles(anime_list):
    print("Anime Titles:")
    for anime in anime_list:
        title = anime.get('anime_title', 'No Title')
        print(title)
    
    # Print the total number of titles
    print(f"\nTotal number of anime titles printed: {len(anime_list)}")

def save_to_excel(anime_list, username):
    wb = Workbook()
    ws = wb.active
    ws.title = "Recently Updated Anime"

    # Adding headers
    headers = ['Anime Title', 'Updated At']
    ws.append(headers)

    # Adding data rows
    for anime in anime_list:
        title = anime.get('anime_title', 'No Title')
        updated_at = anime.get('updated_at', 'No Date')
        updated_at_date = datetime.fromtimestamp(int(updated_at), tz=timezone.utc).strftime('%Y-%m-%d %H:%M:%S') if updated_at != 'No Date' else 'No Date'
        ws.append([title, updated_at_date])

    # Set column width for column A (Anime Title)
    ws.column_dimensions[get_column_letter(1)].width = 125  # Adjust width as needed

    # Save the workbook
    file_name = f'recently_updated_by_{username}.xlsx'
    wb.save(file_name)
    print(f"Data saved to {file_name}")

if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python tracker.py <username> <number_of_months>")
        sys.exit(1)

    username = sys.argv[1]
    months = int(sys.argv[2])

    anime_entries = fetch_all_anime_entries(username)
    filtered_anime = filter_anime_by_date(anime_entries, months)
    print_anime_titles(filtered_anime)
    save_to_excel(filtered_anime, username)
