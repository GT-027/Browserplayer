from flask import Flask, jsonify, send_from_directory, request
import openpyxl
import pandas as pd
import os
import json

app = Flask(__name__)

# Directory where the .webm files are stored
WEBM_DIRECTORY = 'downloaded_webm_files'

def sanitize_filename(name):
    return "".join([c if c.isalnum() or c in " .-_()" else "_" for c in name])

def extract_links_from_excel(file_path, sheet_name, column_name):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet = wb[sheet_name]
    links = []
    
    col_idx = None
    for col in sheet.iter_cols(1, sheet.max_column):
        if col[0].value == column_name:
            col_idx = col[0].column
            break

    if col_idx is None:
        raise ValueError(f"Column '{column_name}' not found in sheet '{sheet_name}'")
    
    for cell in sheet.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2, values_only=False):
        for c in cell:
            if c.hyperlink:
                links.append(c.hyperlink.target)

    return links

def extract_names_from_excel(file_path, sheet_name, columns):
    df = pd.read_excel(file_path, sheet_name=sheet_name)
    
    if not all(col in df.columns for col in columns):
        raise ValueError("One or more specified columns are missing in the Excel sheet.")
    
    combined_names = df[columns[0]].astype(str).str[:30] + " " + \
                     df[columns[1]].astype(str).str[:30] + " - " + \
                     df[columns[2]].astype(str).str[:30]
    full_names = df[columns[0]].astype(str) + " " + df[columns[1]].astype(str) + " - " + df[columns[2]].astype(str)
    
    return combined_names.tolist(), full_names.tolist()

def extract_lyrics(file_path, full_names):
    df = pd.read_excel(file_path)
    
    # Create a dictionary with combined names as keys and lyrics as values
    lyrics_dict = {}
    for i in range(len(df)):
        combined_name = df.iloc[i]['Anime Name'] + " " + df.iloc[i]['Song Type'] + " - " + df.iloc[i]['Song Info']
        lyrics = df.iloc[i]['Lyrics']
        # Treat None or NaN lyrics as empty strings
        if pd.isna(lyrics):
            lyrics = ""
        lyrics_dict[combined_name] = lyrics
    
    # Fetch lyrics for each full name, defaulting to empty string if not found
    lyrics = [lyrics_dict.get(name, "") for name in full_names]
    
    return lyrics

@app.route('/playlist')
def get_playlist():
    amqdb_file_path = './AMQDB.xlsx'
    lyrics_file_path = './Anime_Lyrics.xlsx'
    sheet_name = request.args.get('sheet', 'GaleHail')
    failed = request.args.get('failed', 'false').lower() == 'true'
    column_name = 'Song Info'  # Column containing the links

    try:
        if failed:
            # Load JSON data from file
            with open('amq_song_export.json') as f:
                data = json.load(f)

            # Filter songs with correctGuess as False
            failed_songs = [song for song in data['songs'] if not song['correctGuess']]
            playlist = []
            
            for song in failed_songs:
                song_info = song['songInfo']
                anime_name = song_info['animeNames']['romaji']
                song_name = f"{song_info['songName']} by {song_info['artist']}"

                if song_info['type'] == 1:
                    song_type = f"Opening {song_info['typeNumber']}"
                elif song_info['type'] == 2:
                    song_type = f"Ending {song_info['typeNumber']}"
                elif song_info['type'] == 3:
                    song_type = f"Insert {song_info['typeNumber']}"
                else:
                    song_type = f"Type {song_info['type']}"

                full_name = f"{anime_name} {song_type} - {song_name}"
                sanitized_filename = sanitize_filename(full_name) + ".webm"
                local_file_path = os.path.join(WEBM_DIRECTORY, sanitized_filename)
                if os.path.exists(local_file_path):
                    playlist.append((f"/video/{sanitized_filename}", anime_name, full_name, True, ""))
                else:
                    playlist.append((song['videoUrl'], anime_name, full_name, False, f"Download as {sanitized_filename}"))

            return jsonify(playlist)

        else:
            links = extract_links_from_excel(amqdb_file_path, sheet_name, column_name)
            combined_names, full_names = extract_names_from_excel(amqdb_file_path, sheet_name, ['Anime Name', 'Song Type', 'Song Info'])
            lyrics = extract_lyrics(lyrics_file_path, full_names)
            combined_playlist = []

            for link, combined_name, full_name, lyric in zip(links, combined_names, full_names, lyrics):
                sanitized_filename = sanitize_filename(full_name) + ".webm"
                local_file_path = os.path.join(WEBM_DIRECTORY, sanitized_filename)
                if os.path.exists(local_file_path):
                    combined_playlist.append((f"/video/{sanitized_filename}", combined_name, full_name, True, lyric))
                else:
                    combined_playlist.append((link, combined_name, full_name, False, lyric))
            
            return jsonify(combined_playlist)
    except ValueError as e:
        return jsonify({'error': str(e)}), 400

@app.route('/video/<filename>')
def serve_video(filename):
    return send_from_directory(WEBM_DIRECTORY, filename)

@app.route('/')
def serve_index():
    return send_from_directory('', 'index.html')

if __name__ == '__main__':
    app.run(debug=True)
